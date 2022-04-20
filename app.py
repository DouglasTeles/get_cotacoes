from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import time
from selenium.webdriver.support.select import Select
import openpyxl
import getpass
import pandas as pd
from datetime import datetime
from openpyxl import workbook, load_workbook
import os

print(os.getlogin())

global driver

driver = webdriver.Chrome()

selic = ''
result = 0
bdados = []
bezerros_ro = []

def resetPlanilha():
    tabela = load_workbook("C:\\Users\\"+os.getlogin()+"\\Sicredi\\01.0821 Sede Univales - Documentos\\General\\sidi_relatorios\\Python\\controle\\retornos.xlsx")
    aba_ativa = tabela.active
    aba_ativa["C2"] = 0
    tabela.save("C:\\Users\\"+os.getlogin()+"\\Sicredi\\01.0821 Sede Univales - Documentos\\General\\sidi_relatorios\\Python\\controle\\retornos.xlsx")

def openInfomoney():
    print("Iniciando navegador")
    global driver
    driver.get('https://www.infomoney.com.br/')

def getDolarIbovespa():
    time.sleep(3)
    print("Tentando capturar Valor do Dolar")
    dolar = driver.find_element(By.XPATH, '/html/body/div[4]/div[2]/div[1]/div[3]/div[1]/table/tbody/tr[1]/td[3]')
    dolar = dolar.text
    dolar = dolar.replace("R$", "")
    bdados.append(dolar)

    # trata dolar
    dolar = dolar.replace(",",".")
    dolar = float(dolar)
    dolar = ('Dolar: ', dolar)
    
    print(dolar)

    time.sleep(3)
    print("Tentando capturar o ibovespa")
    ibovespa = driver.find_element(By.XPATH,'/html/body/div[4]/div[2]/div[1]/div[1]/div[1]/div[2]')
    ibovespa = ibovespa.text
    bdados.append(ibovespa)

    # trata ibovespa
    ibovespa = ibovespa.replace(",",".").replace("pts", "")
    ibovespa = float(ibovespa)
    ibovespa = ('Ibovespa: ', ibovespa)
    

    print(ibovespa)
    
def getSojaImea():
    print('Iniciando captura de Soja - IMEA')
    driver.get('https://imea.com.br/imea-site/indicador-soja')
    time.sleep(10)
    soja_imea = driver.find_element(By.XPATH, '/html/body/div[1]/div[3]/section/div[2]/div[1]/div/div[2]/table/tbody/tr[1]/td[2]/small')
    soja_imea = soja_imea.text
    bdados.append(soja_imea)

    # trata SojaImea
    soja_imea = soja_imea.replace(",",".")
    soja_imea = float(soja_imea)
    soja_imea = ('Soja - IMEA: ', soja_imea)
    print(soja_imea)

def getLeite():
    print('Iniciando captura de Leite')
    driver.get('https://imea.com.br/imea-site/indicador-leite')
    time.sleep(10)
    leite = driver.find_element(By.XPATH, '/html/body/div[1]/div[3]/section/div[2]/div[1]/div/div[2]/table/tbody/tr[1]/td[2]/small')
    leite = leite.text
    bdados.append(leite)

    # trata Leite
    leite = leite.replace(",",".")
    leite = float(leite)
    leite = ('Leite: ', leite)
    print(leite)

def getBezerrosMT():
    print('Iniciando captura de Bezerros - MT - IMEA')
    driver.get('https://imea.com.br/imea-site/indicador-boi')
    time.sleep(10)
    bezerro_mt = driver.find_element(By.XPATH, '/html/body/div[1]/div[3]/section/div[2]/div[9]/div/div[2]/table/tbody/tr[1]/td[2]/small')
    bezerro_mt = bezerro_mt.text
    bdados.append(bezerro_mt)

    # trata BezerroMT
    bezerro_mt = bezerro_mt.replace(".","")
    bezerro_mt = bezerro_mt.replace(",",".")
    bezerro_mt = float(bezerro_mt)
    bezerro_mt = ('Bezerro - MT: ', bezerro_mt)
    print(bezerro_mt)

    bezerra_mt = driver.find_element(By.XPATH, '/html/body/div[1]/div[3]/section/div[2]/div[13]/div/div[2]/table/tbody/tr[1]/td[2]/small')
    bezerra_mt = bezerra_mt.text
    bdados.append(bezerra_mt)

    # trata BezerraMT
    bezerra_mt = bezerra_mt.replace(".","")
    bezerra_mt = bezerra_mt.replace(",",".")
    bezerra_mt = float(bezerra_mt)
    bezerra_mt = ('Bezerra - MT: ', bezerra_mt)
    print(bezerra_mt)
    
def getCafe():
    try:
        print('Iniciando captura de cafes')
        driver.get('https://www.noticiasagricolas.com.br/cotacoes/cafe')
        time.sleep(3)
        closeelement = driver.execute_script('document.getElementsByClassName("close")[1].click()')
        time.sleep(3)        
        cafe_Arabica = driver.find_element(By. XPATH,' //div[8]/div/table/tbody/tr[7]/td[2]')
        cafe_Arabica = cafe_Arabica.text
        bdados.append(cafe_Arabica)

        cafe_conilon = driver.find_element(By.XPATH, '//div[12]/div/table/tbody/tr[4]/td[2]')
        cafe_conilon = cafe_conilon.text
        bdados.append(cafe_conilon)

        # Trata Café
        if "." in cafe_Arabica: 
            cafe_Arabica = cafe_Arabica.replace(".","")
            cafe_Arabica = cafe_Arabica.replace(",",".")
            cafe_Arabica = float(cafe_Arabica)
            cafe_Arabica = ('Bezerro - MT: ', cafe_Arabica)
           
            print(cafe_Arabica)
        else:
            cafe_Arabica = cafe_Arabica.replace(",",".")
            cafe_Arabica = float(cafe_Arabica)
            cafe_Arabica = ('Bezerro - MT: ', cafe_Arabica)
            
            print(cafe_Arabica)

    except:
        print("Entrei no exept de mais 3 segundos")
        time.sleep(3)        
        cafe_Arabica = driver.find_element(By. XPATH,' //div[8]/div/table/tbody/tr[7]/td[2]')
        cafe_Arabica = cafe_Arabica.text
        bdados.append(cafe_Arabica)
        print(cafe_Arabica)

        cafe_conilon = driver.find_element(By.XPATH, '//div[12]/div/table/tbody/tr[4]/td[2]')
        cafe_conilon = cafe_conilon.text
        bdados.append(cafe_conilon)
        print(cafe_conilon)

        
def getCacau():
    try:
        print('Iniciando captura de Cacau')
        driver.get('https://www.noticiasagricolas.com.br/cotacoes/cacau')
        time.sleep(5)
        cacau = driver.find_element(By.XPATH, '//div[4]/div[2]/table/tbody/tr/td[2]')
        cacau = cacau.text
        bdados.append(cacau)

        # Trata Cacau
        cacau = cacau.replace(",",".")
        cacau = float(cacau)
        cacau = ('Cacau: ', cacau)
        print(cacau)

        
    except:
        print('Entrei no except. Tentando fechar anuncio')
        driver.get('https://www.noticiasagricolas.com.br/cotacoes/cacau')
        time.sleep(5)
        closeelement = driver.execute_script('document.getElementsByClassName("close")[1].click()')
        time.sleep(3)
        cacau = driver.find_element(By.XPATH, '//div[4]/div[2]/table/tbody/tr/td[2]')
        cacau = cacau.text
        bdados.append(cacau)

        # Trata Cacau
        cacau = cacau.replace(",",".")
        cacau = float(cacau)
        cacau = ('Cacau: ', cacau)
        print(cacau)
      
def getSojaB3():
    try:
        print('Iniciando captura de Soja - B3')
        driver.get('https://www.noticiasagricolas.com.br/cotacoes/soja')
        soja_b3 = driver.find_element(By.XPATH,'//div[1]/div[2]/table/tbody/tr/td[2]')
        soja_b3 = soja_b3.text
        bdados.append(soja_b3)

        # Trata Soja
        soja_b3 = soja_b3.replace(",",".")
        soja_b3 = float(soja_b3)
        soja_b3 = ('Soja - B3: ', soja_b3)
        print(soja_b3)
        
        
    except:
        print('Entrei no except. Tentando fechar anuncio')
        driver.get('https://www.noticiasagricolas.com.br/cotacoes/soja')
        time.sleep(3)
        closeelement = driver.execute_script('document.getElementsByClassName("close")[1].click()')
        time.sleep(2)
        soja_b3 = driver.find_element(By.XPATH,'//div[1]/div[2]/table/tbody/tr/td[2]')
        soja_b3 = soja_b3.text
        bdados.append(soja_b3)

        # Trata Soja
        soja_b3 = soja_b3.replace(",",".")
        soja_b3 = float(soja_b3)
        soja_b3 = ('Cacau: ', soja_b3)
        print(soja_b3)

def getMilhoImea():
    print('Iniciando captura de milho - Imea')
    driver.get('https://imea.com.br/imea-site/indicador-milho')
    time.sleep(10)
    milho_imea = driver.find_element(By.XPATH, '/html/body/div[1]/div[3]/section/div[2]/div[1]/div/div[2]/table/tbody/tr[1]/td[2]/small')
    milho_imea = milho_imea.text
    bdados.append(milho_imea)

    # trata MilhoImea
    milho_imea = milho_imea.replace(",",".")
    milho_imea = float(milho_imea)
    print('Milho - IMEA: ', milho_imea)
    milho_imea = ('Milho - IMEA: ', milho_imea)
    
def getBoiGordo():
    try:
        print('Iniciando captura de Boi gordo')
        driver.get('https://www.noticiasagricolas.com.br/cotacoes/boi')
        time.sleep(5)
        boi_gordo = driver.find_element(By.XPATH, '//*[@class="table-content"]/table/tbody/tr/td[2]')
        boi_gordo = boi_gordo.text
        print('Boi Gordo: ', boi_gordo)
        bdados.append(boi_gordo)

    except:
        print('Entrei no exept')
        driver.get('https://www.noticiasagricolas.com.br/cotacoes/boi')
        time.sleep(3)
        closeelement = driver.execute_script('document.getElementsByClassName("close")[1].click()')
        time.sleep(3)
        boi_gordo = driver.find_element(By.XPATH, '/html/body/div[1]/div[5]/section/div[3]/div[2]/div[1]/div[2]/table/tbody/tr/td[2]')
        print('Boi Gordo: ', boi_gordo.text)
        bdados.append(boi_gordo.text)

def getBezerroRO():
    global bezerro_ro    
    try:
        print('Iniciando captura de Bezerro - RO - b3')
        time.sleep(3)
        bezerro_ro = driver.find_element(By. XPATH, '//div[20]/div[2]/table/tbody/tr[9]/td[2]')
        bezerro_ro = bezerro_ro.text
        bezerros_ro.append(bezerro_ro)
        print(f'Bezerro - RO: {bezerro_ro}')
        
    except:
        print('Entrei no exept')
        driver.get('https://www.noticiasagricolas.com.br/cotacoes/boi')
        time.sleep(3)
        bezerro_ro = driver.find_element(By. XPATH, '//div[20]/div[2]/table/tbody/tr[9]/td[2]')
        bezerro_ro = bezerro_ro.text
        bezerros_ro.append(bezerro_ro)
        print(bezerro_ro) 

        print(f'Bezerro - RO: {bezerro_ro}')
       

def getBezerraRO():
    global bezerra_ro    
    try:
        print('Iniciando captura de Bezerra - RO - b3')
        time.sleep(3)
        bezerra_ro = driver.find_element(By.XPATH, '//div[13]/div/table/tbody/tr[9]/td[2]')
        bezerra_ro = bezerra_ro.text
        bezerros_ro.append(bezerra_ro)
        print(f'Bezerra - RO: {bezerra_ro}')
        
    except:
        print('Entrei no exept')
        driver.get('https://www.noticiasagricolas.com.br/cotacoes/boi')
        time.sleep(3)
         
        bezerra_ro = driver.find_element(By.XPATH, '//div[13]/div/table/tbody/tr[9]/td[2]')
        bezerra_ro = bezerra_ro.text
        bezerros_ro.append(bezerra_ro)
        print(f'Bezerra - RO: {bezerra_ro}')

        
def getMilhoB3():
    try:
        driver.get('https://www.noticiasagricolas.com.br/cotacoes/milho')
        time.sleep(3)
        print('Iniciando captura de milho - B3')
        milho_b3 = driver.find_element(By. XPATH, '//div[1]/div[2]/table/tbody/tr/td[2]')
        milho_b3 = milho_b3.text
        bdados.append(milho_b3)
        
        print('Milho - B3: ', milho_b3)
            
    except:
        print('Entrei no except. Tentando fechar anuncio')
        driver.get('https://www.noticiasagricolas.com.br/cotacoes/milho')
        time.sleep(3)
        closeelement = driver.execute_script('document.getElementsByClassName("close")[1].click()')
        time.sleep(2)
        milho_b3 = driver.find_element(By. XPATH, '//div[1]/div[2]/table/tbody/tr/td[2]')
        milho_b3 = milho_b3.text
        bdados.append(milho_b3)
        print('Milho - B3: ', milho_b3)

def getIPCA():
    ipca = 0
    while ipca == 0:
        print('Iniciando captura de IPCA')  
        driver.get('https://www.ibge.gov.br/explica/inflacao.php')  
        ipca = driver.find_element(By. XPATH, '//*[@id="dadoBrasil"]/li[1]/p[1]')
        print(ipca)
    # ipca = ipca.get_attribute('value')
        ipca = ipca.text
        print(ipca)
        bdados.append(ipca)  
        print ('IPCA: ',ipca)
        ipca = 1
    
    

def getOuro():
    print('Iniciando captura de Valor do Ouro')   
    driver.get('https://www.melhorcambio.com/ouro-hoje')  
    time.sleep(15)
    print("Aguardando ouro")         
    ouro = driver.find_element(By. XPATH, '//*[@id="operacao"]/div/div/input[2]')

    ouro = ouro.get_attribute('value')
    bdados.append(ouro)  
    print ('Ouro: ',ouro)
      
def getSelic():
    print('Iniciando captura de Taxa Selic')
    driver.get('https://www.bcb.gov.br/controleinflacao/historicotaxasjuros')
    time.sleep(5)
    global selic
    selic = ""
    while selic == "": 
        time.sleep(5)
        print("Aguardando selic")
        selic = driver.find_element(By.XPATH, '//*[@id="historicotaxasjuros"]/tbody/tr[1]/td[5]')
        selic = selic.text
    
    bdados.append(selic)
    print("Selic: ",selic)
    selic = selic.replace(",",".")
    selic = float(selic)
    driver.quit()
       
def calcCDI(selic):
    print("Iniciando calculo de CDI")
    cdi_anual = selic - 0.10
    bdados.append(cdi_anual)
    print("CDI Anual: ", cdi_anual)

    cdi_mensal = cdi_anual/12
    
    
    
    cdi_mensal = "{:.2f}".format(cdi_mensal)
    print("CDI mensal: ", cdi_mensal)
    bdados.append(cdi_mensal)

def calcPoupanca(selic):
    print("Iniciando calculo de Poupancas")
    tr = 0
    if selic > 8.5:
        poupanca_anual = 6.17 + tr
        bdados.append(poupanca_anual)
    else:
        poupanca_anual = selic*7/100
        bdados.append(poupanca_anual)

    print("Poupanca Anual: ", poupanca_anual)    

def ultimoUpdate():
    datahora = datetime.today().strftime('%d-%m-%Y %H:%M')
    print(datahora)
    bdados.append(datahora)

def createExcel(bdados, bezerros_ro):
    print("Iniciando gravacao dos dados")
    descricoes = ['Dolar','Ibovespa', 'Milho - IMEA', 'Soja - IMEA',
                 'Bezerro - MT','Bezerra - MT','Leite','Boi Gordo', 'Cafe Arabica', 'Cafe Conilon','Cacau','Soja - B3',
                 'Milho - B3','IPCA','Ouro','Selic', 'CDI Anual','CDI mensal','Poupanca Anual', 'Ultima atualização']
    book = openpyxl.Workbook()
    dados_base =  book['Sheet']
    dados_base.append(descricoes)
    dados_base.append(bdados)
    

    descricoes = ['bezerro_ro', 'bezerra_ro']
    dados_base.append(descricoes)
    dados_base.append(bezerros_ro)

    # "C:\\Users\\"+os.getlogin()+"\\Sicredi\\01.0821 Sede Univales - Documentos\\General\\sidi_relatorios\\Cotacoes\\base.xlsx"
    # destino = "C:\\Users\\"+os.getlogin()+"\\Sicredi\\01.0821 Sede Univales - General\\sidi_relatorios\\Cotacoes\\base.xlsx"
    destino ="C:\\Users\\"+os.getlogin()+"\\Sicredi\\01.0821 Sede Univales - Documentos\\General\\sidi_relatorios\\Cotacoes\\base.xlsx"
    book.save(destino)
    base = pd.read_excel(destino)
    base = base.transpose()
    base.to_excel(destino)


    book = openpyxl.Workbook()
    
   
    
def gravaSucesso():
    print("Indo gravar sucesso")
    tabela = load_workbook("C:\\Users\\"+os.getlogin()+"\\Sicredi\\01.0821 Sede Univales - Documentos\\General\\sidi_relatorios\\Python\\controle\\retornos.xlsx")
    aba_ativa = tabela.active
    aba_ativa["C2"] = 1
    tabela.save("C:\\Users\\"+os.getlogin()+"\\Sicredi\\01.0821 Sede Univales - Documentos\\General\\sidi_relatorios\\Python\\controle\\retornos.xlsx")


def main():
    resetPlanilha()
    openInfomoney()
    getDolarIbovespa()
    getMilhoImea()
    getSojaImea()  
    getBezerrosMT()
    getLeite()
    # # # # # ######Funções tratadas ⬇
    getBoiGordo()
    getBezerroRO()
    getBezerraRO()
    getCafe()
    getCacau()     
    getSojaB3()
    getMilhoB3()
    # Funções tratadas ⬆
    getIPCA()
    getOuro()
    getSelic()    
    calcCDI(selic)
    calcPoupanca(selic)
    ultimoUpdate()
    


main()
createExcel(bdados, bezerros_ro)
gravaSucesso()
driver.quit()

